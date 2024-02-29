
import pickle
import sys
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException
from credenciales import *
import openpyxl
from datetime import date
from datetime import datetime
def obtener_fecha():
    fecha_actual = date.today()
    hora_actual = datetime.now()
    return f"{fecha_actual.day}-{fecha_actual.month}-{fecha_actual.year} --- {hora_actual.hour}.{hora_actual.minute}.{hora_actual.second}"
def iniciar_chrome():
    options = Options()
    user_agent  = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.111 Safari/537.36 OPR/72.0.3815.178"
    options.add_argument(f'user-agent={user_agent}')
    options.add_argument("--disable-notifications")
    options.add_argument("--window-size=1600,800")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--no-sandbox")
    options.add_argument("--log-level=3")
    options.add_argument("--disable-web-security")
    options.add_argument("--no-first-run")
    service_chrome = Service(ChromeDriverManager(path="./chromedriver").install())
    driver = webdriver.Chrome(service=service_chrome, options=options)
    return driver


def iniciar_sesion():
    print("Login en FACEBOOK")
    driver.get(website)
    try:
        element = wait.until(ec.visibility_of_element_located((By.ID, 'email')))
        element.click()
        element.send_keys(user)
    except TimeoutException:
        print("ERROR: No se encontro email")
        return 'ERROR'

    try:
        element = wait.until(ec.visibility_of_element_located((By.ID, 'pass')))
        element.click()
        element.send_keys(password)
    except TimeoutException:
        print("ERROR: No se encontró password")
        return 'ERROR'

    try:
        btn_login = wait.until(ec.element_to_be_clickable((By.XPATH, '//button[@name="login"]')))
        btn_login.click()
    except TimeoutException:
        print("ERROR: No se encontró boton loggin")
        return 'ERROR'

    try:
        wait.until(ec.visibility_of_element_located((By.CSS_SELECTOR, 'div[aria-label="Historias"')))
        # Guardando cookies
        cookies = driver.get_cookies()
        file_cookies = open('utils/facebook.cookies', 'wb')
        pickle.dump(cookies, file_cookies)
        print('Se han guardado las cookies: OK')
        return 'OK'
    except TimeoutException:
        print('ERROR: No se ha cargado las historias')
        return 'ERROR'
def leer_excel():
    wb = openpyxl.load_workbook('./utils/40001-42305 VERIFICAR.xlsx')
    ws = wb['Sheet']
    return ws
def crear_excel(fecha_actual):
    wbNew = openpyxl.Workbook()
    wbNew.save(f'utils/{fecha_actual}.xlsx')
def actualizar_excel(fecha_actual):
    wbNew = openpyxl.load_workbook(f'./utils/{fecha_actual}.xlsx')
    wsNew = wbNew.active
    return wbNew, wsNew
def analizando_datos():
    print("Iniciando el análisis de datos")
    nombreArchivo = '40001-42305 VERIFICADO'
    crear_excel(nombreArchivo)
     ## esto se borra cuando
    ws = leer_excel()
    wbNew, wsNew = actualizar_excel(nombreArchivo)
    max_fila = ws.max_row
    conteoDeMiembros = 0 
    conteoNoMiembros = 0
    conteoVerificaciones = 0
    driver.get('link del grupo')

    for fila in range(1, max_fila+1): 
        url = ws.cell(row=fila, column=2).value 
        nombre = ws.cell(row=fila, column=1).value
        try:
            driver.get(url)
            wait.until(ec.visibility_of_element_located((By.CSS_SELECTOR, "div[class='xieb3on'] span[dir]")))
            text = driver.find_element(By.CSS_SELECTOR, "div[class='xieb3on'] span[dir]")
            if "anterior" in text.text:
                conteoNoMiembros += 1
                print(f'{fila}) Es un miembro anterior ----- X ----- {nombre} - {url}')
            else:
                conteoDeMiembros += 1
                cell1 = wsNew.cell(row=conteoDeMiembros, column=1)
                cell2 = wsNew.cell(row=conteoDeMiembros, column=2)
                cell1.value = nombre
                cell2.value = url
                print(
                    f'{fila}) Es un miembro vigente ----- ✓ ----- {nombre} - {url}')

            if conteoDeMiembros % 10 == 0:
                print("---Guardando datos---")
                wbNew.save(f'./utils/{nombreArchivo}.xlsx')
                wbNew, wsNew = actualizar_excel(nombreArchivo)
            time.sleep(2)

        except TimeoutException:
            try:
                wait.until(
                    ec.visibility_of_element_located((By.CSS_SELECTOR, " a[aria-label='Ir a la sección de noticias']")))
                conteoNoMiembros += 1
                print(f'{fila}) No es miembro----- X ----- {nombre} - {url}')
            except TimeoutException:
                conteoDeMiembros += 1
                conteoVerificaciones += 1
                cell3 = wsNew.cell(row=conteoDeMiembros, column=3)
                cell4 = wsNew.cell(row=conteoDeMiembros, column=4)
                cell3.value = nombre
                cell4.value = url
                print(f'{fila}) Verificar ----- X ----- {nombre} - {url}')

    print("---Guardando datos---")
    print(
        f'Se ha encontrado: \n Miembros:  {conteoDeMiembros}\n No miembros: {conteoNoMiembros}\n '
        f'Total de datos analizados: {conteoDeMiembros + conteoNoMiembros} \n'
        f'Total de datos a verificar: {conteoVerificaciones}')

    wbNew.save(f'./utils/{nombreArchivo}.xlsx')

if __name__ == '__main__':
    print("--Iniciando el programa--")
    start = time.time()
    driver = iniciar_chrome()
    wait = WebDriverWait(driver, 5)
    respuesta = iniciar_sesion()
    if respuesta == "ERROR":
        driver.close()
        sys.exit()
    analizando_datos()
    end = time.time()
    print(f'EL tiempo transcurrido de ejecucion fue aproximadamente {(end - start) / 60} minutos')
input('Pulsa ENTER para finalizar')